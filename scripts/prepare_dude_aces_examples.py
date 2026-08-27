#!/usr/bin/env python3
"""Build the public DUD-E ACES example workbooks used by DockMate-VS."""

from __future__ import annotations

import argparse
import hashlib
import random
import tempfile
import urllib.request
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from rdkit import Chem


TARGET_PAGE = "https://dude.docking.org/targets/aces"
CITATION = "https://doi.org/10.1021/jm300687e"
PDB_PAGE = "https://www.rcsb.org/structure/1E66"
SOURCE_FILES = {
    "actives_final.ism": {
        # The DUD-E server is legacy infrastructure; checksums protect the
        # download when plain HTTP is required for a complete transfer.
        "url": "http://dude.docking.org/targets/aces/actives_final.ism",
        "sha256": "d3c04077ab78028fa104a585b91f8b1e80361d71923f66800a9386866a3d1168",
    },
    "decoys_final.ism": {
        "url": "http://dude.docking.org/targets/aces/decoys_final.ism",
        "sha256": "da06b6b9d0a4ab51c5dfad2938242dc877ef453c67bc433d10568cd88fd1280d",
    },
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _obtain_sources(source_dir: Path) -> None:
    source_dir.mkdir(parents=True, exist_ok=True)
    for name, metadata in SOURCE_FILES.items():
        destination = source_dir / name
        if not destination.is_file() or _sha256(destination) != metadata["sha256"]:
            print(f"Downloading {metadata['url']}")
            urllib.request.urlretrieve(metadata["url"], destination)
        checksum = _sha256(destination)
        if checksum != metadata["sha256"]:
            raise RuntimeError(
                f"Checksum mismatch for {name}: expected {metadata['sha256']}, "
                f"received {checksum}"
            )


def _read_ism(path: Path, active: bool) -> list[dict]:
    records = []
    seen = set()
    expected_fields = 3 if active else 2
    for line_number, line in enumerate(path.read_text().splitlines(), start=1):
        fields = line.split()
        if len(fields) != expected_fields:
            raise ValueError(f"Unexpected {path.name} line {line_number}: {line}")
        smiles = fields[0]
        source_id = fields[2] if active else fields[1]
        if source_id in seen:
            continue
        molecule = Chem.MolFromSmiles(smiles)
        if molecule is None:
            raise ValueError(f"Invalid SMILES in {path.name} line {line_number}")
        seen.add(source_id)
        records.append({
            "source_id": source_id,
            "smiles": smiles,
            "canonical_smiles": Chem.MolToSmiles(
                molecule, canonical=True, isomericSmiles=True
            ),
        })
    return records


def _style_sheet(sheet, widths: dict[str, float]) -> None:
    header_fill = PatternFill("solid", fgColor="177E89")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _add_table(sheet, name: str) -> None:
    table = Table(displayName=name, ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    sheet.add_table(table)


def _write_metadata(workbook: Workbook, rows: Iterable[tuple[str, object]]) -> None:
    sheet = workbook.create_sheet("Metadata")
    sheet.append(("Field", "Value"))
    for row in rows:
        sheet.append(row)
    _style_sheet(sheet, {"A": 27, "B": 105})
    _add_table(sheet, "ExampleMetadata")


def _save_protocol_workbook(destination: Path) -> None:
    workbook = Workbook()
    workbook.properties.title = "DockMate-VS DUD-E ACES protocol-development example"
    workbook.properties.subject = "Native-pose protocol development for PDB 1E66/HUX"
    sheet = workbook.active
    sheet.title = "Docking_Jobs"
    sheet.append((
        "Protein", "Target_Ligand", "PDB_ID", "Ligand", "Chain", "SMILES",
        "label", "Notes",
    ))
    sheet.append((
        "DUD-E ACES", "HUX_native_redock", "1E66", "HUX", "A", None, 1,
        "Native 1E66/HUX pose-recovery control; ligand resolved from RCSB PDB",
    ))
    _style_sheet(
        sheet,
        {"A": 18, "B": 26, "C": 11, "D": 11, "E": 9, "F": 24, "G": 9, "H": 72},
    )
    _add_table(sheet, "ProtocolDevelopmentJobs")
    _write_metadata(workbook, (
        ("Example", "DockMate-VS DUD-E ACES protocol development"),
        ("Purpose", "Exercise native-pose recovery and protocol-factor sweeps"),
        ("Target", "ACES - acetylcholinesterase"),
        ("Receptor", "RCSB PDB 1E66, chain A"),
        ("Co-crystal ligand", "HUX - (-)-huprine X"),
        ("PDB page", PDB_PAGE),
        ("DUD-E target page", TARGET_PAGE),
        ("DUD-E citation", CITATION),
        ("Control semantics", "label=1 identifies the native positive used for protocol development"),
        ("Interpretation", "Software workflow example; exact poses and scores depend on tool versions and hardware"),
    ))
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def _save_screening_workbook(
    destination: Path,
    actives: list[dict],
    decoys: list[dict],
    seed: int,
) -> None:
    workbook = Workbook()
    workbook.properties.title = "DockMate-VS DUD-E ACES enrichment example"
    workbook.properties.subject = "Labelled DUD-E active/decoy screening at PDB 1E66"
    sheet = workbook.active
    sheet.title = "Docking_Jobs"
    sheet.append((
        "Protein", "Target_Ligand", "PDB_ID", "Ligand", "Chain", "SMILES",
        "label", "Source_ID", "Activity_Class", "Canonical_Isomeric_SMILES",
        "Notes",
    ))
    rows = []
    for record in actives:
        rows.append((
            "DUD-E ACES", f"DUD-E_active_{record['source_id']}", "1E66", "HUX", "A",
            record["smiles"], 1, record["source_id"], "active",
            record["canonical_smiles"],
            "DUD-E ACES clustered active; selected independently of docking score",
        ))
    for record in decoys:
        rows.append((
            "DUD-E ACES", f"DUD-E_decoy_{record['source_id']}", "1E66", "HUX", "A",
            record["smiles"], 0, record["source_id"], "matched decoy",
            record["canonical_smiles"],
            "DUD-E ACES property-matched decoy; presumed inactive, not experimentally confirmed",
        ))
    random.Random(seed + 1).shuffle(rows)
    for row in rows:
        sheet.append(row)
    _style_sheet(sheet, {
        "A": 18, "B": 36, "C": 11, "D": 11, "E": 9, "F": 68,
        "G": 9, "H": 22, "I": 20, "J": 68, "K": 72,
    })
    _add_table(sheet, "ScreeningBenchmarkJobs")
    _write_metadata(workbook, (
        ("Example", "DockMate-VS DUD-E ACES screening benchmark"),
        ("Purpose", "Exercise labelled screening, ROC/PR, enrichment, and result charts"),
        ("Target", "ACES - acetylcholinesterase"),
        ("Receptor", "RCSB PDB 1E66, chain A; binding site defined by HUX"),
        ("DUD-E target page", TARGET_PAGE),
        ("DUD-E citation", CITATION),
        ("Active source", "actives_final.ism (DUD-E clustered actives)"),
        ("Active source SHA-256", SOURCE_FILES["actives_final.ism"]["sha256"]),
        ("Decoy source", "decoys_final.ism (DUD-E property-matched decoys)"),
        ("Decoy source SHA-256", SOURCE_FILES["decoys_final.ism"]["sha256"]),
        ("Subset", f"{len(actives)} actives and {len(decoys)} decoys"),
        ("Selection", f"Python random.Random({seed}); sampled after source-ID deduplication; no docking scores used"),
        ("Row ordering", f"Deterministically shuffled with random.Random({seed + 1})"),
        ("Label semantics", "1=DUD-E active; 0=DUD-E matched decoy"),
        ("Decoy caveat", "Matched decoys are presumed non-binders and are not experimentally validated inactives"),
        ("Interpretation", "Demonstrates DockMate-VS functionality; it is not a new scoring-function benchmark"),
        ("Rebuild command", "python scripts/prepare_dude_aces_examples.py"),
    ))
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def build_examples(
    source_dir: Path,
    output_dir: Path,
    seed: int = 42,
    active_count: int = 20,
    decoy_count: int = 200,
) -> None:
    _obtain_sources(source_dir)
    active_records = _read_ism(source_dir / "actives_final.ism", active=True)
    decoy_records = _read_ism(source_dir / "decoys_final.ism", active=False)
    if active_count > len(active_records) or decoy_count > len(decoy_records):
        raise ValueError("Requested subset is larger than the available DUD-E source data")
    generator = random.Random(seed)
    actives = generator.sample(active_records, active_count)
    decoys = generator.sample(decoy_records, decoy_count)
    _save_protocol_workbook(
        output_dir / "dude_aces_protocol_development_1E66.xlsx"
    )
    _save_screening_workbook(
        output_dir / f"dude_aces_screening_subset_1E66_seed{seed}.xlsx",
        actives, decoys, seed,
    )


def main() -> None:
    repository_root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-dir", type=Path)
    parser.add_argument("--output-dir", type=Path, default=repository_root / "examples")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--actives", type=int, default=20)
    parser.add_argument("--decoys", type=int, default=200)
    args = parser.parse_args()
    if args.source_dir:
        build_examples(
            args.source_dir, args.output_dir, args.seed, args.actives, args.decoys
        )
    else:
        with tempfile.TemporaryDirectory(prefix="dockmate_dude_aces_") as temporary:
            build_examples(
                Path(temporary), args.output_dir, args.seed,
                args.actives, args.decoys,
            )


if __name__ == "__main__":
    main()
