#!/usr/bin/env python3
"""
Generate an annotated Excel input template for the Redock Analysis GUI.

The GUI loads its input sheet through ``_load_pairs_from_excel`` in
``docking_platform_gui/gui/redock_analysis.py``. This script produces a
workbook whose *first* sheet is a ready-to-edit input table (the only sheet
the GUI reads) plus several human-only reference sheets that explain every
column and use case.

Every example row below is written so that it actually loads through the GUI
detector. The important behaviours it demonstrates:

* Columns are auto-detected by a normalised name (lower-case, alphanumerics
  only), so ``PDB_ID``/``pdb id``/``PDBCode`` all resolve to the PDB column.
* PDB codes must be 4 characters or the row is silently skipped.
* A row's role is *inferred* in the default (no-label) layout:
    - a row WITH ``decoy SMILES`` -> target becomes an ACTIVE (label 1) and
      each comma-separated decoy becomes a DECOY (label 0) in the same site;
    - a row WITHOUT decoys -> a plain redock/self-dock case (label = None).
* IMPORTANT: adding a ``label``/``class``/``active`` column switches the GUI
  into explicit-label mode and DISABLES decoy auto-expansion. The two layouts
  are therefore mutually exclusive; the primary sheet uses the decoy layout
  and the ``Alt_Explicit_Labels`` sheet shows the alternative.

Usage:
    python scripts/make_gui_input_template.py [--output docking_gui_input_template.xlsx]
"""

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --- shared styling ---------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
BOLD = Font(bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(cell, comment=None):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = BORDER
    if comment:
        c = Comment(comment, "GUI template")
        c.width = 320
        c.height = 200
        cell.comment = c


# --- primary input sheet ----------------------------------------------------

# (header, width, header-comment). The comment is the accepted-alias + meaning
# documentation surfaced when you hover the header cell in Excel.
PRIMARY_COLUMNS = [
    ("PDB_ID", 10,
     "REQUIRED. 4-character PDB code of the receptor structure.\n"
     "Rows whose code is blank or not exactly 4 characters are SILENTLY "
     "SKIPPED by the GUI.\n"
     "Accepted headers: PDB_ID, PDB, PDB code, PDBCode, PDB_code."),
    ("Ligand", 9,
     "Optional. 3-letter residue name (HET code) of the co-crystal ligand.\n"
     "Used to (a) center the docking box on that ligand and (b) provide the "
     "crystal pose for redock RMSD.\n"
     "Leave BLANK for an apo target -> the ligand becomes 'APO' and the site "
     "must come from 'site_residues' or 'pocket_center_*'.\n"
     "Accepted headers: Ligand, resname, ligand_resname, lig."),
    ("Chain", 7,
     "Optional. Chain identifier of the ligand/site (e.g. A).\n"
     "Accepted headers: Chain, ligand_chain."),
    ("Target_Ligand", 18,
     "Optional but recommended. Human-readable name of the compound being "
     "docked; used to label results and output folders.\n"
     "Accepted headers: Target_Ligand, target_ligand_name."),
    ("SMILES", 34,
     "Optional. SMILES of the compound to dock.\n"
     "Used only when 'Use SMILES column' is ticked in the GUI; otherwise the "
     "ligand is taken from the crystal structure.\n"
     "Required for apo / SMILES-only screening rows.\n"
     "Accepted headers: SMILES, smile, smiles_string."),
    ("decoy SMILES", 40,
     "Optional. One or more DECOY SMILES, comma-separated, in a single cell.\n"
     "Presence of any decoy turns this row into an ENRICHMENT set: the target "
     "ligand becomes an ACTIVE (label 1) and every decoy becomes a DECOY "
     "(label 0) docked into the same site.\n"
     "NOTE: this auto-expansion only works when the sheet has NO label/class/"
     "active column.\n"
     "Accepted headers: decoy SMILES, decoy_smiles, decoy_smile."),
    ("decoy compound", 26,
     "Optional. Names for the decoys, comma-separated, PAIRED BY POSITION with "
     "'decoy SMILES' (1st name <-> 1st SMILES, ...).\n"
     "Missing names fall back to decoy_1, decoy_2, ...\n"
     "Accepted headers: decoy compound, decoy_compound, decoy."),
    ("site_residues", 22,
     "Optional. Explicit binding-site residues (e.g. 'A:57, A:102, A:189').\n"
     "Highest-priority site definition: if present, the box is built around "
     "these residues (site_mode = residues).\n"
     "Accepted headers: site_residues, pocket_residues, binding_site_residues, "
     "residues."),
    ("pocket_center_x", 15,
     "Optional. Explicit grid-box CENTER x coordinate (Angstrom).\n"
     "Provide all three of x/y/z to pin the box directly (useful for apo "
     "targets). EXAMPLE values in this template must be replaced with real "
     "coordinates for YOUR structure.\n"
     "Accepted headers: pocket_center_x, site_center_x, grid_center_x, "
     "center_x."),
    ("pocket_center_y", 15,
     "Optional. Explicit grid-box CENTER y coordinate (Angstrom). See "
     "pocket_center_x."),
    ("pocket_center_z", 15,
     "Optional. Explicit grid-box CENTER z coordinate (Angstrom). See "
     "pocket_center_x."),
    ("Notes", 60,
     "Free text. NOT read by the GUI (unrecognised column names are ignored). "
     "Use it for your own annotations."),
]

# Site-definition priority the loader applies (highest first):
#   site_residues  ->  co-crystal Ligand  ->  pocket_center_*  ->  blind/predict

PRIMARY_ROWS = [
    # 1. Redock control / self-docking validation (co-crystal, no decoys)
    dict(PDB_ID="3PTB", Ligand="BEN", Chain="A", Target_Ligand="Benzamidine",
         SMILES="NC(=N)c1ccccc1",
         decoy_SMILES="", decoy_compound="",
         site_residues="", cx="", cy="", cz="",
         Notes="USE CASE - Redock control (self-docking). No decoys, so this "
               "is a plain case (control_label = None). The crystal ligand BEN "
               "is redocked into 3PTB and RMSD vs the crystal pose validates "
               "the protocol. Site = co-crystal ligand BEN."),
    # 2. Enrichment: one active + several decoys, same site
    dict(PDB_ID="1M17", Ligand="AQ4", Chain="A", Target_Ligand="Erlotinib",
         SMILES="C#Cc1cccc(Nc2ncnc3cc(OCCOC)c(OCCOC)cc23)c1",
         decoy_SMILES="CC(=O)Oc1ccccc1C(=O)O, Cn1cnc2c1c(=O)n(C)c(=O)n2C, CC(C)Cc1ccc(C(C)C(=O)O)cc1",
         decoy_compound="aspirin, caffeine, ibuprofen",
         site_residues="", cx="", cy="", cz="",
         Notes="USE CASE - Enrichment set. Target 'Erlotinib' becomes the "
               "ACTIVE (label 1); the three comma-separated decoy SMILES become "
               "DECOYS (label 0), all docked into the EGFR site defined by "
               "co-crystal ligand AQ4. Decoy names pair by position."),
    # 3. SMILES-only screen of an analog in a known pocket (no decoys)
    dict(PDB_ID="1M17", Ligand="AQ4", Chain="A", Target_Ligand="Gefitinib (analog screen)",
         SMILES="COc1cc2ncnc(Nc3ccc(F)c(Cl)c3)c2cc1OCCCN1CCOCC1",
         decoy_SMILES="", decoy_compound="",
         site_residues="", cx="", cy="", cz="",
         Notes="USE CASE - Screen an analog from SMILES. Tick 'Use SMILES "
               "column' so this SMILES (not the crystal ligand) is docked into "
               "the AQ4 site. Plain case (label None). Same PDB as row 2 -> the "
               "structure is reused for multiple compounds."),
    # 4. Apo target, site defined by residues
    dict(PDB_ID="4DFR", Ligand="", Chain="A", Target_Ligand="Methotrexate",
         SMILES="CN(Cc1cnc2nc(N)nc(N)c2n1)c1ccc(C(=O)N[C@@H](CCC(=O)O)C(=O)O)cc1",
         decoy_SMILES="", decoy_compound="",
         site_residues="A:5, A:7, A:27, A:57, A:94",
         cx="", cy="", cz="",
         Notes="USE CASE - Apo / residue-defined site. Ligand is BLANK so the "
               "case is APO; the box is built from 'site_residues' (highest "
               "priority). Dock the SMILES compound. Requires 'Use SMILES "
               "column'. Replace the example residues with real ones."),
    # 5. Explicit grid center (e.g. apo, known box center)
    dict(PDB_ID="1STP", Ligand="", Chain="A", Target_Ligand="Biotin",
         SMILES="O=C(O)CCCC[C@@H]1SC[C@@H]2NC(=O)N[C@H]12",
         decoy_SMILES="", decoy_compound="",
         site_residues="",
         cx=20.5, cy=25.0, cz=32.0,
         Notes="USE CASE - Explicit grid center. Site pinned by "
               "pocket_center_x/y/z (Angstrom). EXAMPLE COORDINATES - replace "
               "with the real box center for 1STP before running."),
    # 6. Cofactor / additive filter demonstration
    dict(PDB_ID="1A6M", Ligand="HEM", Chain="A", Target_Ligand="Heme (cofactor demo)",
         SMILES="",
         decoy_SMILES="", decoy_compound="",
         site_residues="", cx="", cy="", cz="",
         Notes="USE CASE - Additive/cofactor filter. HEM is a known COFACTOR. "
               "With 'Exclude cofactors' ticked this row is DROPPED, unless it "
               "is a labelled control and 'Always include controls' is on. "
               "Shows how the exclude filters interact with controls."),
]


def build_primary_sheet(ws):
    ws.title = "Docking_Input"

    headers = [c[0] for c in PRIMARY_COLUMNS]
    for col_idx, (name, width, comment) in enumerate(PRIMARY_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        _style_header(cell, comment)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for r, row in enumerate(PRIMARY_ROWS, start=2):
        values = [
            row["PDB_ID"], row["Ligand"], row["Chain"], row["Target_Ligand"],
            row["SMILES"], row["decoy_SMILES"], row["decoy_compound"],
            row["site_residues"], row["cx"], row["cy"], row["cz"], row["Notes"],
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if headers[c - 1] == "Notes":
                cell.fill = NOTE_FILL

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    # A couple of cell-level teaching comments on the first enrichment row.
    def _cell_comment(coord, text):
        c = Comment(text, "GUI template")
        c.width = 300
        c.height = 160
        ws[coord].comment = c

    _cell_comment(
        "F3",
        "Multiple decoys go in ONE cell, comma-separated. Presence of decoys "
        "makes the target an ACTIVE and each of these a DECOY, all docked into "
        "the same site for enrichment (ROC-AUC / EF).",
    )
    _cell_comment(
        "G3",
        "Decoy names pair by position with 'decoy SMILES': aspirin<->1st, "
        "caffeine<->2nd, ibuprofen<->3rd.",
    )


# --- reference sheets -------------------------------------------------------

def _title_row(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24


def _table_header(ws, row, headers, widths):
    for col_idx, (name, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.fill = SUBHEAD_FILL
        cell.font = BOLD
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _table_body(ws, start_row, rows):
    for i, values in enumerate(rows):
        r = start_row + i
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP_TOP
            cell.border = BORDER


def build_use_cases_sheet(ws):
    ws.title = "Use_Cases"
    _title_row(ws, "What each row of 'Docking_Input' demonstrates", 4)
    headers = ["Row", "Use case", "How to express it", "What the GUI does"]
    widths = [6, 26, 46, 60]
    _table_header(ws, 3, headers, widths)
    rows = [
        ("1", "Redock control (self-docking)",
         "PDB + co-crystal Ligand, no decoys.",
         "Redocks the crystal ligand and reports RMSD vs the crystal pose. "
         "control_label = None. This is the 'redock control' / protocol "
         "validation case."),
        ("2", "Enrichment (active vs decoys)",
         "PDB + Ligand + Target_Ligand + one-or-more 'decoy SMILES' "
         "(comma-separated) with matching 'decoy compound' names.",
         "Target becomes ACTIVE (label 1); each decoy becomes DECOY (label 0), "
         "all docked into the same site. Enables ROC-AUC / enrichment factors."),
        ("3", "SMILES analog screen",
         "PDB + Ligand + SMILES, no decoys; tick 'Use SMILES column'.",
         "Docks the provided SMILES (not the crystal ligand) into the site "
         "defined by the co-crystal ligand. control_label = None."),
        ("4", "Apo target, residue-defined site",
         "PDB, blank Ligand, 'site_residues' filled, SMILES to dock.",
         "Ligand = APO; box built from residues (site_mode = residues). "
         "Highest-priority site definition."),
        ("5", "Apo target, explicit box center",
         "PDB, blank Ligand, pocket_center_x/y/z filled.",
         "Box pinned to the given XYZ center. Use when you know the pocket "
         "location. Replace example coordinates with real ones."),
        ("6", "Additive / cofactor filtering",
         "A row whose Ligand is a cofactor (HEM) or crystallisation additive "
         "(GOL, EDO, SO4 ...).",
         "Dropped when 'Exclude cofactors'/'Exclude additives' is on, unless it "
         "is a labelled control and 'Always include controls' is on."),
    ]
    _table_body(ws, 4, rows)

    note_row = 4 + len(rows) + 1
    ws.cell(row=note_row, column=1,
            value="Site-definition priority (highest first):").font = BOLD
    ws.cell(row=note_row + 1, column=1,
            value="site_residues  ->  co-crystal Ligand  ->  pocket_center_x/y/z  ->  blind/prediction")
    ws.cell(row=note_row + 3, column=1,
            value="Key gotcha:").font = BOLD
    ws.cell(row=note_row + 4, column=1,
            value="Decoy auto-expansion (row 2) only works when the sheet has NO "
                  "label/class/active column. Add such a column and the GUI "
                  "switches to explicit-label mode (see 'Alt_Explicit_Labels').")


def build_column_reference_sheet(ws):
    ws.title = "Column_Reference"
    _title_row(ws, "Recognised columns (auto-detected by normalised name)", 4)
    headers = ["Column (as used here)", "Required?", "Accepted header aliases", "Meaning"]
    widths = [22, 12, 40, 60]
    _table_header(ws, 3, headers, widths)
    rows = [
        ("PDB_ID", "Yes",
         "PDB_ID, PDB, PDB code, PDBCode, PDB_code",
         "4-char PDB code of the receptor. Rows without a valid 4-char code "
         "are silently skipped."),
        ("Ligand", "No*",
         "Ligand, resname, ligand_resname, lig",
         "Co-crystal ligand HET code; defines the site and crystal pose. Blank "
         "-> APO. *Falls back to the 2nd column if unnamed and not using SMILES."),
        ("Chain", "No", "Chain, ligand_chain",
         "Chain identifier for the ligand/site."),
        ("Target_Ligand", "No",
         "Target_Ligand, target_ligand_name",
         "Human-readable compound name used for labels/output folders."),
        ("SMILES", "No",
         "SMILES, smile, smiles_string",
         "Compound to dock when 'Use SMILES column' is on; required for "
         "apo/SMILES-only rows."),
        ("decoy SMILES", "No",
         "decoy SMILES, decoy_smiles, decoy_smile",
         "Comma-separated decoy SMILES in one cell; presence triggers the "
         "active+decoys enrichment expansion."),
        ("decoy compound", "No",
         "decoy compound, decoy_compound, decoy",
         "Comma-separated decoy names, paired by position with 'decoy SMILES'."),
        ("site_residues", "No",
         "site_residues, pocket_residues, binding_site_residues, residues",
         "Explicit binding-site residues, e.g. 'A:57, A:102'. Highest-priority "
         "site definition."),
        ("pocket_center_x/y/z", "No",
         "pocket_center_x, site_center_x, grid_center_x, center_x (and y/z)",
         "Explicit grid-box center in Angstrom; provide all three."),
        ("label / class", "No (special)",
         "label, class, is_active, active, actives",
         "If present, switches to explicit-label mode (1/active vs 0/decoy) and "
         "DISABLES decoy auto-expansion. See 'Alt_Explicit_Labels'."),
        ("(any other column)", "-", "-",
         "Ignored by the GUI. Safe to use for your own notes."),
    ]
    _table_body(ws, 4, rows)


def build_alt_labels_sheet(ws):
    ws.title = "Alt_Explicit_Labels"
    _title_row(ws, "ALTERNATIVE layout: explicit active/decoy labels", 6)
    ws.cell(row=2, column=1,
            value="Use this layout INSTEAD of the decoy-SMILES layout when you "
                  "want to label each molecule explicitly (one molecule per "
                  "row). Adding a 'label' column disables decoy auto-expansion. "
                  "To use it in the GUI, make this the first sheet.").alignment = WRAP_TOP
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 42

    headers = ["PDB_ID", "Ligand", "Target_Ligand", "SMILES", "label", "Notes"]
    widths = [10, 9, 22, 40, 8, 46]
    _table_header(ws, 4, headers, widths)
    rows = [
        ("1M17", "AQ4", "Erlotinib",
         "C#Cc1cccc(Nc2ncnc3cc(OCCOC)c(OCCOC)cc23)c1", 1,
         "Explicit ACTIVE. label accepts 1/true/yes/active/positive."),
        ("1M17", "AQ4", "aspirin",
         "CC(=O)Oc1ccccc1C(=O)O", 0,
         "Explicit DECOY. label accepts 0/false/no/decoy/inactive/negative."),
        ("1M17", "AQ4", "caffeine",
         "Cn1cnc2c1c(=O)n(C)c(=O)n2C", 0,
         "Another decoy, docked into the same AQ4 site."),
        ("3PTB", "BEN", "Benzamidine",
         "NC(=N)c1ccccc1", "",
         "Blank label -> plain redock case (control_label = None)."),
    ]
    _table_body(ws, 5, rows)
    for i in range(len(rows)):
        ws.cell(row=5 + i, column=6).fill = NOTE_FILL


def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--output", default="docking_gui_input_template.xlsx",
                        help="Output .xlsx path (default: docking_gui_input_template.xlsx)")
    args = parser.parse_args()

    wb = Workbook()
    build_primary_sheet(wb.active)
    build_use_cases_sheet(wb.create_sheet("Use_Cases"))
    build_column_reference_sheet(wb.create_sheet("Column_Reference"))
    build_alt_labels_sheet(wb.create_sheet("Alt_Explicit_Labels"))

    out = Path(args.output)
    wb.save(out)
    print(f"Wrote {out.resolve()} ({out.stat().st_size} bytes)")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
